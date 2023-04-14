#!/usr/bin/env python3
#-*- coding:UTF-8 -*-

#dslogic的I2C解析结果的csv文件,将一列数据的格式转换成一行一个数据包的格式

import sys
import os

try:
    import csv
    import openpyxl
except ImportError:
    print ("找不到依赖库");
    print (ImportError);

if len (sys.argv) > 1:
    str_file_path = sys.argv[1]
else:
    str_file_path = input ("请输入文件路径：")

if os.path.isfile(str_file_path):
    str_file_dir = os.path.dirname(str_file_path);
    str_file_base = os.path.basename(str_file_path);
    str_file_base = os.path.splitext(str_file_base)[0];
    if str_file_dir == "":
        str_fileout_path = os.path.join(".", str_file_base + ".xlsx"); # 兼容linux和windows路径
    else:
        str_fileout_path = os.path.join(str_file_dir, str_file_base + ".xlsx");

    with open (str_file_path, "r") as file_sheet:
        csv_sheet_reader = csv.reader(file_sheet, delimiter=',');

        # 将csv文件经过预处理去掉第一行和第一列并转换成列表
        list_csv_data = [];
        for index, item in enumerate(csv_sheet_reader):
            if (index == 0):
                continue;
            list_csv_data.append(item[2]);

        # 删掉无意义的符号
        list_data = [];
        for index, item in enumerate(list_csv_data):
            data_temp = item;

            if ((data_temp == "ACK") or (data_temp == "NACK") or (data_temp == "Stop")):
                continue;

            if (data_temp == "Start repeat"):
                data_temp = "Start";

            if (data_temp.find(':') != -1):
                data_temp = data_temp.split(":")[1]; # 以':'为界定符分割文本
                data_temp = data_temp.strip(" "); # 去除字符串中的空格，即':'后面可能存在的空格

            # data_temp = "{:02X}".format(data_temp); # 修改数字格式

            list_data.append(data_temp);

        # 将一维列表转换成二维列表
        list_sheet = [];
        int_sheet_row = -1;
        for item in list_data:
            if (item == "Start"):
                int_sheet_row = int_sheet_row + 1;
                list_sheet.append([item]);
            else:
                list_sheet[int_sheet_row].extend([item]);

    #生成xlsx文件
    workbook = openpyxl.Workbook()
    sheet = workbook.active #Excel页签标识

    for list_data_row in list_sheet:
        sheet.append(list_data_row);

    #设置列宽
    for index in range(1, 2):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = 6;
    for index in range(3, sheet.max_column+1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = 3;

    #保存为xlsx文件，名字可以随意写
    workbook.save(str_fileout_path)

    '''
    # 保存为csv文件
    with open (str_fileout_path, "w", newline="") as file_result: # newline="" 是为了之后csv.writerow没有多余的空行
        file_result.truncate(); #清空文件
        csv_sheet_writer = csv.writer(file_result, delimiter=',');

        for item in list_sheet:
            csv_sheet_writer.writerow(item);
    '''

elif os.path.isdir(str_file_path):
    print ("传入参数的路径无效")

else:
    print ("传入参数未知错误")

exit();
