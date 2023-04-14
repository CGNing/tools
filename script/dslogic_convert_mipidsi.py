#!/usr/bin/env python3
#-*- coding:UTF-8 -*-

#dslogic的MIPI解析结果的csv文件,将一列数据的格式转换成一行一个数据包的格式

import sys
import os

try:
    import csv
    import openpyxl
    import openpyxl.utils
    import openpyxl.styles
except ImportError:
    print ("找不到依赖库");
    print (ImportError);

def calc_ecc():
    return

def calc_crc():
    return

if len (sys.argv) > 1:
    str_file_path = sys.argv[1]
else:
    str_file_path = input ("请输入文件路径：")

if os.path.isfile(str_file_path):
    str_file_dir = os.path.dirname(str_file_path);
    str_file_base = os.path.basename(str_file_path);
    str_file_base = os.path.splitext(str_file_base)[0];
    if str_file_dir == "":
        str_fileout_path = os.path.join(".", str_file_base + ".xlsx"); # 兼容linux和windows路径
    else:
        str_fileout_path = os.path.join(str_file_dir, str_file_base + ".xlsx");

    with open (str_file_path, "r") as file_sheet:
        csv_sheet_reader = csv.reader(file_sheet, delimiter=',');

        # 将csv文件经过预处理去掉第一行和第一列并转换成列表
        list_csv_data = [];
        for index, item in enumerate(csv_sheet_reader):
            if (index == 0):
                continue;
            list_csv_data.append(item[2]);

        # 删掉无意义的符号
        list_data = [];
        for index, item in enumerate(list_csv_data):
            data_temp = item;

            if (data_temp == "Escape mode entry"):
                data_temp = "ESC";
            elif (data_temp == "Bi-directional Data Lane Turnaround"):
                data_temp = "BTA";

            list_data.append(data_temp);

        # 将一维列表转换成二维列表
        list_sheet = [["Mode", "Length", "Action", "DT", "Data0", "Data1", "ECC"]];
        int_sheet_row = 1;
        bool_newline = True;

        for item in list_data:
            if (item == "Stop"):
                int_sheet_row = int_sheet_row + 1;
                bool_newline = True;
            elif (bool_newline == True):
                bool_newline = False;
                list_sheet.append([item]);
            else:
                list_sheet[int_sheet_row].extend([item]);

    list_sheet_property = [[0]];
    mipi_data_direct = 0;   # 0 host -> slv, 1 slv -> host
    for index in range(1, len(list_sheet)):
        #插入MIPI包的长度
        len_list_sheet_index = len(list_sheet[index])-1;
        if (len_list_sheet_index > 0):
            list_sheet[index].insert(1, "0x{:02X}".format(len_list_sheet_index));

        #判断数据包的方向
        if (list_sheet[index][0] == "BTA"):
            mipi_data_direct = not(mipi_data_direct);
            list_sheet_property.append([2]);
        else:
            list_sheet_property.append([mipi_data_direct]);

    # 生成xlsx文件
    workbook = openpyxl.Workbook()
    sheet = workbook.active #Excel页签标识

    for list_sheet_row in list_sheet:
        sheet.append(list_sheet_row);

    #设置列宽
    for index in range(1, sheet.max_column+1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = 6;\

    #设置颜色
    fill_cambridgeblue = openpyxl.styles.PatternFill("solid", "DAEEF3");
    fill_blue = openpyxl.styles.PatternFill("solid", "B7DEE8");

    ##第一行粗体
    for cell in sheet[1]:
                cell.font=openpyxl.styles.Font(bold=True);

    for index in range(0, len(list_sheet)):
        if (list_sheet_property[index][0] == 2):
            for cell in sheet[index+1]:
                cell.fill=fill_blue;
        elif (list_sheet_property[index][0] == 1):
            for cell in sheet[index+1]:
                cell.fill=fill_cambridgeblue;

    #保存为xlsx文件，名字可以随意写
    workbook.save(str_fileout_path)

    '''
    # 保存为csv文件
    with open (str_fileout_path, "w", newline="") as file_result: # newline="" 是为了之后csv.writerow没有多余的空行
        file_result.truncate(); #清空文件
        csv_sheet_writer = csv.writer(file_result, delimiter=',');

        for item in list_sheet:
            csv_sheet_writer.writerow(item);
    '''

elif os.path.isdir(str_file_path):
    print ("传入参数的路径无效")

else:
    print ("传入参数未知错误")

exit();
